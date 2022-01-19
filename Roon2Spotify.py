from envparse import env
import pprint
import sys
import urllib.parse
import argparse
import logging
from functools import partial
from openpyxl import load_workbook
import spotipy
import spotipy.util as util

env.read_envfile()
pprint.pprint (env.str)
my_client_id = env('SPOTIPY_CLIENT_ID')
my_client_secret = env('SPOTIPY_CLIENT_SECRET')
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger()
spotify_token=""
# Define retry util function
# via Stack Overflow https://stackoverflow.com/questions/21786382/pythonic-way-of-retry-running-a-function
# Define Exception class for retry
class RetryException(Exception):
    u_str = "Exception ({}) raised after {} tries."

    def __init__(self, exp, max_retry):
        self.exp = exp
        self.max_retry = max_retry    
    def __unicode__(self):
        return self.u_str.format(self.exp, self.max_retry)
    def __str__(self):
        return self.__unicode__()
        
def retry_func(func, max_retry=10):
    """
    @param func: The function that needs to be retry
    @param max_retry: Maximum retry of `func` function, default is `10`
    @return: func
    @raise: RetryException if retries exceeded than max_retry
    """
    for retry in range(1, max_retry + 1):
        try:
            return func()
        except Exception as e:
            logger.info('Failed to call {}, in retry({}/{})'.format(func.func,
                                                           retry, max_retry))
    else:
        raise RetryException('Failed to call function after retrying.', max_retry)
    
def find_spotify(artist, album_name):
    sp = spotipy.Spotify(auth=spotify_token)
    print ("searching for {}:{}".format(artist, album_name))
    search = "artist:{0} \"{1}]\"".format(artist, album_name).encode('utf-8')
    result = retry_func(partial(sp.search, q=search, type='album', limit=1), max_retry=1)
    items = result['albums']['items']
    try:
        id = items[0]['id']
        retry_func(partial(sp.current_user_saved_albums_add, albums=[id,]), max_retry=1)
    except IndexError:
        print ("Error search for {} {}".format(artist, album_name))
    except RetryException as e:
        print (e.args)
def populate_albums(file):
    wb = load_workbook(filename = file)
    sheet = wb.active
    max_row = sheet.max_row
    current_album = ""
    last_album = ""
    for i in range(2, max_row+1):
        current_album = sheet.cell(i, 2).value
        artist = sheet.cell(i,1).value
        if current_album != last_album:
            find_spotify(artist, current_album)
            last_album = current_album

def populate_playlists(file):
    pass
def flush_library():
    logger.info("Flushing Library")
    sp = spotipy.Spotify(auth=spotify_token)
    
    results = sp.current_user_saved_albums()
    albums = results['items']
    for album in albums:
        tracks = sp.album_tracks(album['id'])
        sp.current_user_saved_tracks_delete(tracks)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Populate Spotify from your Roon Library.")
    parser.add_argument('-a', '--album', dest='albums')
    parser.add_argument('-p', '--playlist', dest='playlists' )
    parser.add_argument('--flush', default=False, action="store_true")
    file = "test_file.xlsx"
    spotify_token = util.prompt_for_user_token("JoeCotellese",
                                    "playlist-modify-private,user-library-modify,user-library-read",
                                    client_id=my_client_id,
                                    client_secret=my_client_secret,
                                    redirect_uri='http://localhost')

    args = parser.parse_args()
    if args.flush == True:
        flush_library()
    if args.albums:
        populate_albums(args.albums)
